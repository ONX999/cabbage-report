"""
輸出格式化模組 - 負責產生Excel和PDF報表
"""

import os
from datetime import datetime
from typing import Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class OutputFormatter:
    """
    處理報表輸出格式化的類別
    """

    def __init__(self, output_dir: str = "output"):
        self.output_dir = output_dir
        self._ensure_output_dir()

    def _ensure_output_dir(self):
        """確保輸出目錄存在"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def export_to_excel(self, data: pd.DataFrame, statistics: Dict,
                       insights: List[str], summary_table: pd.DataFrame,
                       filename: str = None) -> str:
        """
        匯出資料到Excel檔案

        Args:
            data: 原始價格資料
            statistics: 統計資料
            insights: 洞察分析
            summary_table: 摘要統計表
            filename: 檔案名稱

        Returns:
            str: 產生的檔案路徑
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"甘藍價格報表_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()

        # 移除預設工作表
        wb.remove(wb.active)

        # 建立摘要工作表
        self._create_summary_sheet(wb, statistics, insights)

        # 建立統計表工作表
        if not summary_table.empty:
            self._create_statistics_sheet(wb, summary_table)

        # 建立原始資料工作表
        if not data.empty:
            self._create_data_sheet(wb, data)

        wb.save(filepath)
        return filepath

    def _create_summary_sheet(self, wb: Workbook, statistics: Dict, insights: List[str]):
        """建立摘要工作表"""
        ws = wb.create_sheet("摘要報告", 0)

        # 標題
        ws['A1'] = "甘藍價格分析報告"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')

        # 產生時間
        ws['A3'] = "報告產生時間:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 統計資料
        row = 5
        ws[f'A{row}'] = "統計摘要"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        for key, value in statistics.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1

        # 洞察分析
        row += 2
        ws[f'A{row}'] = "重要發現"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        for insight in insights:
            ws[f'A{row}'] = f"• {insight}"
            row += 1

        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

    def _create_statistics_sheet(self, wb: Workbook, summary_table: pd.DataFrame):
        """建立統計表工作表"""
        ws = wb.create_sheet("市場統計")

        # 標題
        ws['A1'] = "市場別統計表"
        ws['A1'].font = Font(size=14, bold=True)

        # 插入資料
        for r in dataframe_to_rows(summary_table, index=True, header=True):
            ws.append(r)

        # 格式化標題行
        for cell in ws[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # 調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_data_sheet(self, wb: Workbook, data: pd.DataFrame):
        """建立原始資料工作表"""
        ws = wb.create_sheet("原始資料")

        # 標題
        ws['A1'] = "原始交易資料"
        ws['A1'].font = Font(size=14, bold=True)

        # 插入資料
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)

        # 格式化標題行
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # 調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_to_pdf(self, statistics: Dict, insights: List[str],
                     summary_table: pd.DataFrame, filename: str = None) -> str:
        """
        匯出資料到PDF檔案

        Args:
            statistics: 統計資料
            insights: 洞察分析
            summary_table: 摘要統計表
            filename: 檔案名稱

        Returns:
            str: 產生的檔案路徑
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"甘藍價格報表_{timestamp}.pdf"

        filepath = os.path.join(self.output_dir, filename)

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []

        # 樣式設定
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )

        # 標題
        title = Paragraph("甘藍價格分析報告", title_style)
        story.append(title)

        # 產生時間
        timestamp_text = f"報告產生時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        story.append(Paragraph(timestamp_text, styles['Normal']))
        story.append(Spacer(1, 20))

        # 統計摘要
        story.append(Paragraph("統計摘要", styles['Heading2']))
        for key, value in statistics.items():
            text = f"<b>{key}:</b> {value}"
            story.append(Paragraph(text, styles['Normal']))
        story.append(Spacer(1, 20))

        # 重要發現
        story.append(Paragraph("重要發現", styles['Heading2']))
        for insight in insights:
            story.append(Paragraph(f"• {insight}", styles['Normal']))
        story.append(Spacer(1, 20))

        # 市場統計表
        if not summary_table.empty:
            story.append(Paragraph("市場別統計表", styles['Heading2']))

            # 準備表格資料
            table_data = [['市場名稱'] + list(summary_table.columns)]
            for idx, row in summary_table.iterrows():
                table_data.append([str(idx)] + [str(val) for val in row])

            # 建立表格
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)

        doc.build(story)
        return filepath