"""
報表匯出模組 - 負責將分析結果匯出為Excel與PDF格式
"""

import logging
from datetime import datetime
from typing import Dict, List
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel報表匯出器"""

    def __init__(self):
        """初始化Excel匯出器"""
        self.workbook = None
        self.styles = self._create_styles()

    def export_report(self,
                     data: pd.DataFrame,
                     statistics: Dict,
                     insights: List[str],
                     output_path: str,
                     include_charts: bool = True) -> str:
        """
        匯出完整報表到Excel檔案

        Args:
            data: 價格資料DataFrame
            statistics: 統計分析結果
            insights: 洞察分析結果
            output_path: 輸出檔案路徑
            include_charts: 是否包含圖表

        Returns:
            str: 輸出檔案路徑
        """
        try:
            self.workbook = Workbook()

            # 移除預設工作表
            self.workbook.remove(self.workbook.active)

            # 建立各個工作表
            self._create_summary_sheet(statistics, insights)
            self._create_data_sheet(data)
            self._create_statistics_sheet(statistics)

            if include_charts and not data.empty:
                self._create_charts_sheet(data)

            # 儲存檔案
            self.workbook.save(output_path)
            logger.info("Excel報表已匯出至: %s", output_path)

            return output_path

        except Exception as e:
            logger.error("匯出Excel報表時發生錯誤: %s", str(e))
            raise

    def _create_styles(self) -> Dict:
        """建立樣式定義"""
        return {
            'header_font': Font(name='Microsoft YaHei', size=14, bold=True, color='FFFFFF'),
            'title_font': Font(name='Microsoft YaHei', size=12, bold=True),
            'normal_font': Font(name='Microsoft YaHei', size=10),
            'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'accent_fill': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'center_align': Alignment(horizontal='center', vertical='center'),
            'left_align': Alignment(horizontal='left', vertical='center')
        }

    def _create_summary_sheet(self, statistics: Dict, insights: List[str]) -> None:
        """建立摘要工作表"""
        ws = self.workbook.create_sheet(title="報表摘要")

        # 標題
        ws['A1'] = "甘藍價格分析報表"
        ws['A1'].font = Font(name='Microsoft YaHei', size=16, bold=True)
        ws.merge_cells('A1:D1')

        # 產生時間
        ws['A3'] = "報表產生時間:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # 基本統計摘要
        row = 5
        ws[f'A{row}'] = "基本統計資料"
        ws[f'A{row}'].font = self.styles['title_font']
        row += 1

        if 'basic_stats' in statistics:
            basic_stats = statistics['basic_stats']
            stats_data = [
                ('平均價格', f"${basic_stats.get('avg_price', 0):.2f}"),
                ('最高價格', f"${basic_stats.get('max_price', 0):.2f}"),
                ('最低價格', f"${basic_stats.get('min_price', 0):.2f}"),
                ('標準差', f"${basic_stats.get('price_std', 0):.2f}"),
                ('資料筆數', f"{basic_stats.get('data_points', 0):,}"),
                ('涵蓋市場數', f"{basic_stats.get('unique_markets', 0)}")
            ]

            for label, value in stats_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1

        # 重要洞察
        row += 2
        ws[f'A{row}'] = "重要洞察"
        ws[f'A{row}'].font = self.styles['title_font']
        row += 1

        for i, insight in enumerate(insights[:10], 1):  # 最多顯示10個洞察
            ws[f'A{row}'] = f"{i}. {insight}"
            ws[f'A{row}'].alignment = self.styles['left_align']
            row += 1

        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25

    def _create_data_sheet(self, data: pd.DataFrame) -> None:
        """建立原始資料工作表"""
        ws = self.workbook.create_sheet(title="原始資料")

        if data.empty:
            ws['A1'] = "無資料"
            return

        # 寫入標題行
        for col_idx, column in enumerate(data.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = self.styles['header_font']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']
            cell.alignment = self.styles['center_align']

        # 寫入資料
        for row_idx, row_data in enumerate(data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.styles['normal_font']
                cell.border = self.styles['border']

                # 數值欄位置中對齊
                if isinstance(value, (int, float)):
                    cell.alignment = self.styles['center_align']
                else:
                    cell.alignment = self.styles['left_align']

        # 調整欄寬
        for col_idx, column in enumerate(data.columns, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 15

    def _create_statistics_sheet(self, statistics: Dict) -> None:
        """建立統計分析工作表"""
        ws = self.workbook.create_sheet(title="統計分析")

        row = 1

        for category, stats in statistics.items():
            if not isinstance(stats, dict):
                continue

            # 類別標題
            ws[f'A{row}'] = category.replace('_', ' ').title()
            ws[f'A{row}'].font = self.styles['title_font']
            ws.merge_cells(f'A{row}:B{row}')
            row += 1

            # 統計項目
            for key, value in stats.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                if isinstance(value, dict):
                    ws[f'B{row}'] = str(value)
                elif isinstance(value, float):
                    ws[f'B{row}'] = f"{value:.2f}"
                else:
                    ws[f'B{row}'] = str(value)
                row += 1

            row += 1  # 空行分隔

        # 調整欄寬
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_charts_sheet(self, data: pd.DataFrame) -> None:
        """建立圖表工作表"""
        ws = self.workbook.create_sheet(title="價格趨勢圖")

        try:
            # 準備圖表資料
            if 'date' in data.columns and 'avg_price' in data.columns:
                # 按日期分組計算平均價格
                daily_prices = data.groupby('date')['avg_price'].mean().reset_index()
                daily_prices = daily_prices.sort_values('date')

                # 將資料寫入工作表
                ws['A1'] = "日期"
                ws['B1'] = "平均價格"

                for i, (_, row) in enumerate(daily_prices.iterrows(), 2):
                    ws[f'A{i}'] = row['date']
                    ws[f'B{i}'] = row['avg_price']

                # 建立折線圖
                chart = LineChart()
                chart.title = "甘藍價格趨勢"
                chart.style = 13
                chart.y_axis.title = '價格 ($)'
                chart.x_axis.title = '日期'

                # 設定資料範圍
                data_range = Reference(ws, min_col=2, min_row=1, max_row=len(daily_prices) + 1)
                chart.add_data(data_range, titles_from_data=True)

                # 設定X軸標籤
                dates_range = Reference(ws, min_col=1, min_row=2, max_row=len(daily_prices) + 1)
                chart.set_categories(dates_range)

                # 新增圖表到工作表
                ws.add_chart(chart, "D2")

        except Exception as e:
            logger.warning("建立圖表時發生錯誤: %s", str(e))
            ws['A1'] = "圖表建立失敗"


class PDFExporter:
    """PDF報表匯出器"""

    def __init__(self):
        """初始化PDF匯出器"""
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def export_report(self,
                     data: pd.DataFrame,
                     statistics: Dict,
                     insights: List[str],
                     output_path: str) -> str:
        """
        匯出完整報表到PDF檔案

        Args:
            data: 價格資料DataFrame
            statistics: 統計分析結果
            insights: 洞察分析結果
            output_path: 輸出檔案路徑

        Returns:
            str: 輸出檔案路徑
        """
        try:
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            story = []

            # 建立報表內容
            story.extend(self._create_header())
            story.extend(self._create_summary_section(statistics))
            story.extend(self._create_insights_section(insights))
            story.extend(self._create_data_summary_section(data))

            # 產生PDF
            doc.build(story)
            logger.info("PDF報表已匯出至: %s", output_path)

            return output_path

        except Exception as e:
            logger.error("匯出PDF報表時發生錯誤: %s", str(e))
            raise

    def _setup_custom_styles(self) -> None:
        """設定自訂樣式"""
        # 標題樣式
        self.styles.add(ParagraphStyle(
            name='ChineseTitle',
            parent=self.styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            spaceAfter=20,
            alignment=1  # 置中
        ))

        # 副標題樣式
        self.styles.add(ParagraphStyle(
            name='ChineseHeading2',
            parent=self.styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10
        ))

        # 內文樣式
        self.styles.add(ParagraphStyle(
            name='ChineseNormal',
            parent=self.styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            spaceAfter=6
        ))

    def _create_header(self) -> List:
        """建立報表標題"""
        elements = []

        # 主標題
        title = Paragraph("甘藍價格分析報表", self.styles['ChineseTitle'])
        elements.append(title)

        # 產生時間
        generation_time = datetime.now().strftime('%Y年%m月%d日 %H:%M')
        time_para = Paragraph(f"報表產生時間：{generation_time}", self.styles['ChineseNormal'])
        elements.append(time_para)
        elements.append(Spacer(1, 20))

        return elements

    def _create_summary_section(self, statistics: Dict) -> List:
        """建立統計摘要區段"""
        elements = []

        # 區段標題
        heading = Paragraph("統計摘要", self.styles['ChineseHeading2'])
        elements.append(heading)

        if 'basic_stats' in statistics:
            basic_stats = statistics['basic_stats']

            # 建立統計表格
            table_data = [
                ['統計項目', '數值'],
                ['平均價格', f"${basic_stats.get('avg_price', 0):.2f}"],
                ['最高價格', f"${basic_stats.get('max_price', 0):.2f}"],
                ['最低價格', f"${basic_stats.get('min_price', 0):.2f}"],
                ['標準差', f"${basic_stats.get('price_std', 0):.2f}"],
                ['資料筆數', f"{basic_stats.get('data_points', 0):,}"],
                ['涵蓋市場數', f"{basic_stats.get('unique_markets', 0)}"]
            ]

            table = Table(table_data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(table)
            elements.append(Spacer(1, 20))

        return elements

    def _create_insights_section(self, insights: List[str]) -> List:
        """建立洞察分析區段"""
        elements = []

        # 區段標題
        heading = Paragraph("重要洞察", self.styles['ChineseHeading2'])
        elements.append(heading)

        # 洞察列表
        for i, insight in enumerate(insights[:10], 1):  # 最多顯示10個洞察
            insight_para = Paragraph(f"{i}. {insight}", self.styles['ChineseNormal'])
            elements.append(insight_para)

        elements.append(Spacer(1, 20))

        return elements

    def _create_data_summary_section(self, data: pd.DataFrame) -> List:
        """建立資料摘要區段"""
        elements = []

        # 區段標題
        heading = Paragraph("資料摘要", self.styles['ChineseHeading2'])
        elements.append(heading)

        if not data.empty:
            # 資料概覽
            summary_text = f"""
            本報表分析了共計 {len(data):,} 筆甘藍價格資料。
            資料涵蓋期間從 {data['date'].min().strftime('%Y-%m-%d') if 'date' in data.columns else '未知'}
            至 {data['date'].max().strftime('%Y-%m-%d') if 'date' in data.columns else '未知'}。
            """

            summary_para = Paragraph(summary_text, self.styles['ChineseNormal'])
            elements.append(summary_para)

            # 如果資料量不大，顯示部分資料樣本
            if len(data) <= 20 and not data.empty:
                sample_heading = Paragraph("資料樣本", self.styles['ChineseHeading2'])
                elements.append(sample_heading)

                # 準備表格資料
                display_columns = ['date', 'market_name', 'avg_price', 'volume']
                available_columns = [col for col in display_columns if col in data.columns]

                if available_columns:
                    table_data = [available_columns]  # 標題行

                    for _, row in data.head(10).iterrows():
                        row_data = []
                        for col in available_columns:
                            value = row[col]
                            if isinstance(value, float):
                                row_data.append(f"{value:.2f}")
                            else:
                                row_data.append(str(value))
                        table_data.append(row_data)

                    # 建立表格
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))

                    elements.append(table)
        else:
            no_data_para = Paragraph("本期間無資料", self.styles['ChineseNormal'])
            elements.append(no_data_para)

        return elements


def export_report(data: pd.DataFrame,
                 statistics: Dict,
                 insights: List[str],
                 output_path: str,
                 format_type: str = 'excel',
                 include_charts: bool = True) -> str:
    """
    匯出報表

    Args:
        data: 價格資料DataFrame
        statistics: 統計分析結果
        insights: 洞察分析結果
        output_path: 輸出檔案路徑
        format_type: 輸出格式 ('excel' 或 'pdf')
        include_charts: 是否包含圖表 (僅Excel適用)

    Returns:
        str: 輸出檔案路徑

    Raises:
        ValueError: 當格式類型不支援時
        Exception: 當匯出失敗時
    """
    format_type = format_type.lower()

    if format_type == 'excel':
        exporter = ExcelExporter()
        return exporter.export_report(data, statistics, insights, output_path, include_charts)
    elif format_type == 'pdf':
        exporter = PDFExporter()
        return exporter.export_report(data, statistics, insights, output_path)
    else:
        raise ValueError(f"不支援的輸出格式: {format_type}")